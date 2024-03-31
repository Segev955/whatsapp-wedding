from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from twilio.rest import Client
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)


def send_initial_message():
    # Function to send the initial message asking if you're coming to the wedding
    host_number = 'whatsapp:+12097526051'  # Replace this with your actual WhatsApp number in E.164 format
    twilio_account_sid = 'AC2def0040777e5b9e17aa21c7f9cbb33b'
    twilio_auth_token = '1b12440ce8ffebbb2ea56c8d811933f1'

    client = Client(twilio_account_sid, twilio_auth_token)
    to_number = 'whatsapp:+972542040149'  # Replace this with your actual phone number in E.164 format

    message_body = "Meital and Segev are getting married 👩‍❤‍👨"
    message_body += "\nAre you coming to the wedding? 🎉"
    message_body += "\n1. Yes"
    message_body += "\n2. Yes, after the canopy"
    message_body += "\n3. No"

    message_body_heb = "מיטל ושגב מתחתנים 👩‍❤‍👨 "
    message_body_heb += "\n אז תגיעו? 🎉"
    message_body_heb += "\n1. בטחחחח"
    message_body_heb += "\n2. אחרי החופה"
    message_body_heb += "\n3. לא יכולים להגיע"

    # Create media URL for the image
    media_url = 'https://pouch.jumpshare.com/preview/wPOjLY_PG-E_g_6M7-w_B5_l7f45ZjHHmwsMSI-R0GH2NlWyIqI6NIcJYaFmIe7sjKYjLLXfqaOzQHQ4ZCX1iRI5wThRYj6kSgOmCvToA0k'  # Replace with the URL of your image

    client.messages.create(body=message_body, from_=host_number, to=to_number, media_url=media_url)


def save_to_excel(phone_number, response):
    # Function to save phone number and response to an Excel file
    excel_file = 'responses.xlsx'

    # Create a new workbook if the file doesn't exist
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(['Phone Number', 'Yes', 'Yes, after the canopy', 'No'])
    else:
        # Load existing workbook
        wb = load_workbook(excel_file)
        ws = wb.active

    # Find if the phone number already exists in the workbook
    existing_row_index = None
    for row_index in range(2, ws.max_row + 1):
        if ws.cell(row=row_index, column=1).value == phone_number:
            existing_row_index = row_index
            break

    # If the phone number already exists, update the response
    if existing_row_index:
        if response == 'yes':
            ws.cell(row=existing_row_index, column=2).value = 'V'
            ws.cell(row=existing_row_index, column=3).value = ''
            ws.cell(row=existing_row_index, column=4).value = ''
        elif response == 'yes, after the canopy':
            ws.cell(row=existing_row_index, column=2).value = ''
            ws.cell(row=existing_row_index, column=3).value = 'V'
            ws.cell(row=existing_row_index, column=4).value = ''
        elif response == 'no':
            ws.cell(row=existing_row_index, column=2).value = ''
            ws.cell(row=existing_row_index, column=3).value = ''
            ws.cell(row=existing_row_index, column=4).value = 'V'
    else:
        # Append phone number and response to the worksheet
        new_row = [phone_number, '', '', '']
        if response == 'yes':
            new_row[1] = 'V'
        elif response == 'yes, after the canopy':
            new_row[2] = 'V'
        elif response == 'no':
            new_row[3] = 'V'
        ws.append(new_row)

    # Save the workbook
    wb.save(excel_file)


@app.route("/webhook", methods=['POST'])
def webhook():
    # Get the incoming message
    incoming_msg = request.values.get('Body', '').lower()

    # Get the sender's phone number
    sender_number = request.values.get('From', '')

    # Create a Twilio response object
    resp = MessagingResponse()
    if incoming_msg in ['1', '2', '3']:
        if incoming_msg == '1':
            incoming_msg = 'yes'
        elif incoming_msg == '2':
            incoming_msg = 'yes, after the canopy'
        elif incoming_msg == '3':
            incoming_msg = 'no'
    # Save the sender's response to Excel
    if incoming_msg.lower() in ['yes', 'yes, after the canopy', 'no']:
        save_to_excel(sender_number, incoming_msg)
        if incoming_msg in ["yes", "yes, after the canopy"]:
            resp.message("Perfect, will see you there :) 🎉")
        elif incoming_msg == 'no':
            resp.message("Ok :(, still love you 🥲")

    else:
        resp.message("Wrong answer: \n1. Yes\n2. Yes, after the canopy\n3. No")

    resp.message("You can change your answer any time")

    return str(resp)


if __name__ == "__main__":
    # Send the initial message when the server starts
    send_initial_message()

    # Run the Flask web server
    port = int(os.environ.get("PORT", 8000))
    app.run(host='0.0.0.0', port=port)


